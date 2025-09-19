import io
import re
from datetime import date, datetime
from typing import Any, Dict, List, Optional, Tuple

from flask import session

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - handled gracefully at runtime
    load_workbook = None

from models import Report, SATReport


_ALIAS_SANITIZE = re.compile(r'[^a-z0-9]+')
_COLLAPSE_WHITESPACE = re.compile(r'\s+')
EMAIL_PATTERN = re.compile(r'^[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}$', re.IGNORECASE)
NAME_PATTERN = re.compile(r"^[A-Za-z][A-Za-z\s\.\-']+$")
CLIENT_PATTERN = re.compile(r"^[A-Za-z0-9][A-Za-z0-9\s\.\-&']+$")
PROJECT_REFERENCE_PATTERN = re.compile(r'^[A-Z0-9][A-Z0-9_\-./ ]{2,}$')
UUID_PATTERN = re.compile(r'([0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12})', re.IGNORECASE)


def _normalize_alias(value: str) -> str:
    if not value:
        return ''
    return _ALIAS_SANITIZE.sub(' ', value.strip().lower()).strip()


def _collapse_whitespace(value: str) -> str:
    return _COLLAPSE_WHITESPACE.sub(' ', value).strip()


def _coerce_to_string(value: Any) -> str:
    if value is None:
        return ''
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, (datetime, date)):
        return value.strftime('%Y-%m-%d')
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value).strip()
    return str(value).strip()


def _normalize_project_reference(value: str) -> str:
    return _collapse_whitespace(value).upper()


def _normalize_email(value: str) -> str:
    return value.lower()


def _has_value(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        return bool(value.strip())
    return True


CONVERSATION_FLOW: List[Dict[str, Any]] = [
    {
        "name": "DOCUMENT_TITLE",
        "label": "Document Title",
        "prompt": "Let's start with the SAT document title.",
        "help_text": "Use the exact wording that should appear on the generated report.",
        "required": True,
        "min_length": 3,
        "max_length": 120,
        "aliases": ("document title", "title", "sat title"),
        "normalizer": _collapse_whitespace,
    },
    {
        "name": "CLIENT_NAME",
        "label": "Client Name",
        "prompt": "Who is the client for this report?",
        "help_text": "Enter the organisation or project owner receiving the SAT.",
        "required": True,
        "min_length": 2,
        "max_length": 120,
        "aliases": ("client", "client name", "customer", "company"),
        "normalizer": _collapse_whitespace,
        "pattern": CLIENT_PATTERN,
        "pattern_error": "Client name can include letters, numbers, spaces and -&'.",
    },
    {
        "name": "PROJECT_REFERENCE",
        "label": "Project Reference",
        "prompt": "What is the project reference or identifier?",
        "help_text": "Provide the internal or client reference code used for this SAT.",
        "required": True,
        "min_length": 3,
        "max_length": 60,
        "aliases": ("project reference", "project id", "project", "reference"),
        "normalizer": _normalize_project_reference,
        "pattern": PROJECT_REFERENCE_PATTERN,
        "pattern_error": "Project reference should be at least 3 characters and use letters, numbers or -_/ .",
    },
    {
        "name": "PURPOSE",
        "label": "Purpose",
        "prompt": "Briefly, what is the purpose of this SAT?",
        "help_text": "Summarise why this acceptance test is being conducted.",
        "required": True,
        "min_length": 10,
        "min_words": 3,
        "max_length": 400,
        "aliases": ("purpose", "objective", "aim", "report purpose"),
        "normalizer": _collapse_whitespace,
    },
    {
        "name": "SCOPE",
        "label": "Scope",
        "prompt": "Summarise the scope of the testing.",
        "help_text": "List the major systems or functions covered by the SAT.",
        "required": True,
        "min_length": 10,
        "min_words": 3,
        "max_length": 600,
        "aliases": ("scope", "testing scope", "scope of work"),
        "normalizer": _collapse_whitespace,
    },
]

ADDITIONAL_FIELDS: List[Dict[str, Any]] = [
    {
        "name": "PREPARED_BY",
        "label": "Prepared By",
        "required": False,
        "min_length": 3,
        "max_length": 80,
        "aliases": ("prepared by", "author", "compiled by"),
        "normalizer": _collapse_whitespace,
        "pattern": NAME_PATTERN,
        "pattern_error": "Prepared by should only include alphabetic characters and punctuation.",
    },
    {
        "name": "USER_EMAIL",
        "label": "Prepared By Email",
        "required": False,
        "aliases": ("prepared by email", "author email", "email", "contact email"),
        "pattern": EMAIL_PATTERN,
        "pattern_error": "Please provide a valid email address (e.g. engineer@example.com).",
        "normalizer": _normalize_email,
    },
    {
        "name": "DOCUMENT_REFERENCE",
        "label": "Document Reference",
        "required": False,
        "min_length": 3,
        "max_length": 60,
        "aliases": ("document reference", "doc reference", "reference number", "ref"),
        "normalizer": _normalize_project_reference,
    },
    {
        "name": "REVISION",
        "label": "Revision",
        "required": False,
        "min_length": 1,
        "max_length": 10,
        "aliases": ("revision", "rev"),
        "normalizer": _normalize_project_reference,
    },
]

FIELD_DEFINITIONS: Dict[str, Dict[str, Any]] = {
    item["name"]: dict(item) for item in CONVERSATION_FLOW
}
for item in ADDITIONAL_FIELDS:
    FIELD_DEFINITIONS[item["name"]] = dict(item)

CONVERSATION_ORDER = [item["name"] for item in CONVERSATION_FLOW]
REQUIRED_FIELDS = {name for name, meta in FIELD_DEFINITIONS.items() if meta.get("required")}

_FIELD_ALIAS_LOOKUP: Dict[str, str] = {}
for field_name, meta in FIELD_DEFINITIONS.items():
    for candidate in {
        field_name,
        meta.get("label", ''),
        *meta.get("aliases", ()),
    }:
        normalized = _normalize_alias(candidate)
        if normalized and normalized not in _FIELD_ALIAS_LOOKUP:
            _FIELD_ALIAS_LOOKUP[normalized] = field_name


class BotConversationState:
    """Encapsulates bot state stored in the user session."""

    SESSION_KEY = "bot_conversation_state"

    def __init__(self) -> None:
        self.position = 0
        self.answers: Dict[str, str] = {}
        self.extracted: Dict[str, str] = {}

    @classmethod
    def load(cls) -> "BotConversationState":
        raw = session.get(cls.SESSION_KEY)
        state = cls()
        if not raw:
            return state
        state.position = raw.get("position", 0)
        state.answers = raw.get("answers", {})
        state.extracted = raw.get("extracted", {})
        return state

    def save(self) -> None:
        session[self.SESSION_KEY] = {
            "position": self.position,
            "answers": self.answers,
            "extracted": self.extracted,
        }

    def reset(self) -> None:
        session.pop(self.SESSION_KEY, None)
        self.position = 0
        self.answers = {}
        self.extracted = {}

    def sync_to_next_question(self) -> None:
        """Advance pointer to the next unanswered field."""
        for idx, field_name in enumerate(CONVERSATION_ORDER):
            value = self.answers.get(field_name) or self.extracted.get(field_name)
            if _has_value(value):
                continue
            self.position = idx
            break
        else:
            self.position = len(CONVERSATION_ORDER)


def start_conversation() -> Dict[str, Any]:
    state = BotConversationState()
    state.reset()
    payload = _build_question_payload(state)
    state.save()
    return payload


def process_user_message(message: str) -> Dict[str, Any]:
    state = BotConversationState.load()

    command_payload = _handle_command(message, state)
    if command_payload is not None:
        state.save()
        return command_payload

    state.sync_to_next_question()
    if state.position >= len(CONVERSATION_ORDER):
        payload = _build_question_payload(state)
        state.save()
        return payload

    field_name = CONVERSATION_ORDER[state.position]
    ok, normalized, error = _apply_validation(field_name, message)
    if not ok:
        meta = FIELD_DEFINITIONS[field_name]
        response: Dict[str, Any] = {
            "completed": False,
            "field": field_name,
            "question": meta["prompt"],
            "collected": _merge_results(state),
            "errors": [error],
            "pending_fields": _pending_fields(state),
        }
        if meta.get("help_text"):
            response["help_text"] = meta["help_text"]
        state.save()
        return response

    if _has_value(normalized):
        state.answers[field_name] = normalized

    state.position += 1
    payload = _build_question_payload(state)
    state.save()
    return payload


def ingest_excel(file_storage) -> Dict[str, Any]:
    state = BotConversationState.load()
    try:
        extracted, warnings = _extract_excel_fields(file_storage)
    except Exception as exc:  # noqa: BLE001 - user facing feedback preferred here
        payload = _build_question_payload(state)
        state.save()

        response: Dict[str, Any] = {
            "error": f"Failed to read Excel file: {exc}",
            "collected": payload.get("collected", _merge_results(state)),
            "completed": payload["completed"],
            "pending_fields": payload.get("pending_fields", []),
        }

        if not payload["completed"]:
            response["field"] = payload["field"]
            response["question"] = payload["question"]
            if "help_text" in payload:
                response["help_text"] = payload["help_text"]

        return response

    if extracted:
        state.extracted.update(extracted)

    message = (
        f"Processed {len(extracted)} fields from {file_storage.filename}."
        if extracted
        else f"No recognised fields found in {file_storage.filename}."
    )

    payload = _build_question_payload(state)
    state.save()

    response: Dict[str, Any] = {
        "message": message,
        "collected": payload.get("collected", _merge_results(state)),
        "completed": payload["completed"],
        "pending_fields": payload.get("pending_fields", []),
    }

    if not payload["completed"]:
        response.update({
            "field": payload["field"],
            "question": payload["question"],
        })
        if "help_text" in payload:
            response["help_text"] = payload["help_text"]

    if warnings:
        response["warnings"] = warnings

    return response


def _extract_excel_fields(file_storage) -> Tuple[Dict[str, str], List[str]]:
    if load_workbook is None:
        raise RuntimeError('Excel support requires the openpyxl package to be installed on the server.')

    content = file_storage.read()
    workbook = load_workbook(filename=io.BytesIO(content), data_only=True)
    file_storage.stream.seek(0)

    extracted: Dict[str, str] = {}
    warnings: List[str] = []

    for sheet in workbook.worksheets:
        sheet_data, sheet_warnings = _extract_fields_from_sheet(sheet)
        for field_name, value in sheet_data.items():
            if field_name in extracted:
                warnings.append(f"{sheet.title}: duplicate value for {_field_label(field_name)} ignored.")
                continue
            extracted[field_name] = value
        warnings.extend(sheet_warnings)

    return extracted, warnings


def _extract_fields_from_sheet(sheet) -> Tuple[Dict[str, str], List[str]]:
    results: Dict[str, str] = {}
    warnings: List[str] = []

    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if not row:
            continue
        for col_idx, cell in enumerate(row):
            header = _coerce_to_string(cell)
            if not header:
                continue
            field_name = _match_field_alias(header)
            if not field_name or field_name in results:
                continue

            value = _extract_candidate_value(sheet, row_idx, col_idx, row)
            if value is None:
                continue

            ok, normalized, error = _apply_validation(field_name, value)
            if not ok:
                warnings.append(
                    f"{sheet.title}: {_field_label(field_name)} at row {row_idx} skipped - {error}"
                )
                continue

            if _has_value(normalized):
                results[field_name] = normalized

    return results, warnings


def _extract_candidate_value(sheet, row_idx: int, col_idx: int, row) -> Optional[str]:
    # Prefer same-row values, fall back to a short lookahead in the same column.
    for candidate in row[col_idx + 1:]:
        candidate_value = _coerce_to_string(candidate)
        if candidate_value:
            return candidate_value

    max_lookahead = 3
    for offset in range(1, max_lookahead + 1):
        target_row = row_idx + offset
        if target_row > sheet.max_row:
            break
        candidate_value = _coerce_to_string(sheet.cell(row=target_row, column=col_idx + 1).value)
        if candidate_value:
            return candidate_value

    return None


def _match_field_alias(header: str) -> Optional[str]:
    normalized = _normalize_alias(header)
    if not normalized:
        return None
    return _FIELD_ALIAS_LOOKUP.get(normalized)


def _field_label(field_name: str) -> str:
    meta = FIELD_DEFINITIONS.get(field_name, {})
    return meta.get("label", field_name.replace('_', ' ').title())


def _apply_validation(field_name: str, raw_value: Any) -> Tuple[bool, Optional[str], Optional[str]]:
    meta = FIELD_DEFINITIONS[field_name]
    label = meta.get("label", field_name.replace('_', ' ').title())
    value = _coerce_to_string(raw_value)

    if not value:
        if meta.get("required"):
            return False, None, f"{label} is required."
        return True, "", None

    if meta.get("strip", True):
        value = value.strip()

    min_length = meta.get("min_length")
    if min_length and len(value) < min_length:
        return False, None, f"{label} must be at least {min_length} characters long."

    min_words = meta.get("min_words")
    if min_words and len(value.split()) < min_words:
        return False, None, f"{label} should contain at least {min_words} words."

    max_length = meta.get("max_length")
    if max_length and len(value) > max_length:
        value = value[:max_length].strip()

    pattern = meta.get("pattern")
    if pattern and not pattern.fullmatch(value):
        return False, None, meta.get("pattern_error", f"{label} has an invalid format.")

    normalizer = meta.get("normalizer")
    if normalizer:
        value = normalizer(value)

    validator = meta.get("validator")
    if validator:
        try:
            value = validator(value, meta)
        except ValueError as exc:
            return False, None, str(exc)

    return True, value, None


def _merge_results(state: BotConversationState) -> Dict[str, str]:
    merged: Dict[str, str] = {}
    combined = dict(state.extracted)
    combined.update(state.answers)
    for key, value in combined.items():
        if _has_value(value):
            merged[key] = value
    return merged


def _pending_fields(state: BotConversationState) -> List[str]:
    pending: List[str] = []
    merged = _merge_results(state)
    for field_name in CONVERSATION_ORDER:
        if not _has_value(merged.get(field_name)):
            pending.append(field_name)
    return pending


def _build_question_payload(state: BotConversationState) -> Dict[str, Any]:
    state.sync_to_next_question()
    merged = _merge_results(state)
    pending = _pending_fields(state)

    if state.position >= len(CONVERSATION_ORDER):
        return {
            "completed": True,
            "collected": merged,
            "pending_fields": pending,
        }

    field_name = CONVERSATION_ORDER[state.position]
    meta = FIELD_DEFINITIONS[field_name]
    payload: Dict[str, Any] = {
        "completed": False,
        "field": field_name,
        "question": meta["prompt"],
        "collected": merged,
        "pending_fields": pending,
    }

    if meta.get("help_text"):
        payload["help_text"] = meta["help_text"]

    return payload


def _parse_document_request(message: str) -> Optional[str]:
    lowered = message.lower().strip()
    if not lowered:
        return None
    if not any(keyword in lowered for keyword in ("download", "fetch", "get")):
        return None
    match = UUID_PATTERN.search(message)
    if match:
        return match.group(1)
    return None


def _handle_command(message: str, state: BotConversationState) -> Optional[Dict[str, Any]]:
    submission_id = _parse_document_request(message)
    if not submission_id:
        return None

    result = resolve_report_download_url(submission_id)
    payload = _build_question_payload(state)

    command: Dict[str, Any] = {
        'type': 'document_fetch',
        'requested_id': submission_id,
        'success': 'download_url' in result,
    }

    if command['success']:
        command['download_url'] = result['download_url']
        metadata = {
            key: result[key]
            for key in ('document_title', 'client_name', 'project_reference')
            if result.get(key)
        }
        if metadata:
            command['metadata'] = metadata
        title = metadata.get('document_title') if metadata else None
        if not title:
            title = 'SAT Report'
        payload.setdefault('messages', []).append(
            f"{title} is ready to download."
        )
    else:
        command['error'] = result.get('error', 'Document not found.')
        payload.setdefault('command_errors', []).append(command['error'])
        payload.setdefault('messages', []).append(command['error'])

    payload['command'] = command
    return payload


def resolve_report_download_url(submission_id: str) -> Dict[str, str]:
    report = Report.query.filter_by(id=submission_id).first()
    if not report:
        return {"error": "Report not found."}
    sat_report = SATReport.query.filter_by(report_id=submission_id).first()
    if not sat_report:
        return {"error": "SAT data not available for this report."}
    return {
        "download_url": f"/status/download-modern/{submission_id}",
        "document_title": report.document_title or "SAT Report",
        "client_name": report.client_name,
        "project_reference": report.project_reference,
    }
